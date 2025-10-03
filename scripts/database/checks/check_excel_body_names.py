from openpyxl import load_workbook

excel_file = 'app/data/All_Materials_Combined_ready_matched_cleaned_density_locale.xlsx'

print("Checking Excel file for lowercase body designations...\n")

wb = load_workbook(excel_file, read_only=True)
print(f"Sheets found: {wb.sheetnames}\n")

total_issues = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Find body_name column
    headers = [cell.value for cell in ws[1]]
    if 'body_name' not in headers:
        print(f"Sheet '{sheet_name}': No 'body_name' column")
        continue
    
    body_name_col = headers.index('body_name')
    
    print(f"Sheet '{sheet_name}':")
    
    # Collect unique body names with issues
    issues = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        body_name = row[body_name_col]
        if body_name and ' Ring' in str(body_name):
            parts = str(body_name).split()
            if len(parts) >= 2 and parts[-1] == 'Ring':
                body_parts = parts[:-2]
                
                for part in body_parts:
                    if len(part) == 1 and part.islower():
                        issues.add(body_name)
                        break
    
    if issues:
        print(f"  ❌ Found {len(issues)} unique body names with lowercase:")
        for name in sorted(list(issues))[:10]:
            print(f"    - {name}")
        if len(issues) > 10:
            print(f"    ... and {len(issues) - 10} more")
        total_issues += len(issues)
    else:
        print(f"  ✅ No lowercase issues")
    print()

wb.close()
print(f"Total unique body names with lowercase across all sheets: {total_issues}")
